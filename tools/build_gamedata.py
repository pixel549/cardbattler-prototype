#!/usr/bin/env python3
import json
import sys
from typing import Any, Dict, List, Optional, Tuple
from openpyxl import load_workbook


def die(msg: str) -> None:
    raise SystemExit(f"[build_gamedata] {msg}")


def norm(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def as_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or norm(v) == "":
        return default
    try:
        return int(float(v))
    except Exception:
        die(f"Expected int, got {v!r}")


def as_float(v: Any, default: Optional[float] = None) -> Optional[float]:
    if v is None or norm(v) == "":
        return default
    try:
        return float(v)
    except Exception:
        die(f"Expected float, got {v!r}")


def as_bool(v: Any, default: bool = False) -> bool:
    if v is None or norm(v) == "":
        return default
    s = norm(v).lower()
    if s in ("true", "1", "yes", "y"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    die(f"Expected bool, got {v!r}")


def split_csv(v: Any) -> List[str]:
    s = norm(v)
    if not s:
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def parse_json_cell(v: Any, default: Any) -> Any:
    s = norm(v)
    if not s:
        return default
    try:
        return json.loads(s)
    except Exception as e:
        die(f"Invalid JSON in cell: {s[:120]}... ({e})")


def sheet_rows_by_header(ws) -> Tuple[List[str], List[Dict[str, Any]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [norm(c) for c in rows[0]]
    if not any(header):
        die(f"Sheet {ws.title!r} missing header row")

    idx = {h: i for i, h in enumerate(header) if h}
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        if all(norm(c) == "" for c in r):
            continue
        obj = {}
        for h, i in idx.items():
            obj[h] = r[i] if i < len(r) else None
        out.append(obj)
    return header, out


def require_fields(row: Dict[str, Any], fields: List[str], ctx: str) -> None:
    missing = [f for f in fields if norm(row.get(f)) == ""]
    if missing:
        die(f"{ctx}: missing required fields: {missing}")


def build_cards(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    cards: Dict[str, Any] = {}
    for row in rows:
        require_fields(row, ["id", "name", "type", "costRAM", "effects_json"], "Cards")
        cid = norm(row["id"])
        if cid in cards:
            die(f"Cards: duplicate id {cid}")

        tags = split_csv(row.get("tags"))
        effects = parse_json_cell(row.get("effects_json"), default=[])
        if not isinstance(effects, list):
            die(f"Cards {cid}: effects_json must be a JSON array")

        trigger = as_float(row.get("mutation_triggerChance"), None)
        tier_weights = {}
        for t in list("ABCDEFGHI"):
            k = f"mutation_tier_{t}"
            w = as_float(row.get(k), None)
            if w is not None:
                tier_weights[t] = w

        mutationOdds = None
        if trigger is not None or tier_weights:
            mutationOdds = {
                "triggerChance": trigger if trigger is not None else 0.25,
                "tiers": tier_weights if tier_weights else {"A": 1},
            }

        brick_w = as_float(row.get("final_brickWeight"), 1.0)
        rewrite_w = as_float(row.get("final_rewriteWeight"), 1.0)
        rewrite_pool = split_csv(row.get("final_rewritePool"))
        brick_behavior = norm(row.get("final_brickBehavior")) or "Exhaust"
        if brick_behavior not in ("Exhaust", "RemoveFromRun"):
            die(f"Cards {cid}: final_brickBehavior must be Exhaust or RemoveFromRun")

        card = {
            "id": cid,
            "name": norm(row["name"]),
            "type": norm(row["type"]),
            "costRAM": as_int(row["costRAM"], 0),
            "effects": effects,
        }

        if tags:
            card["tags"] = tags
        duc = as_int(row.get("defaultUseCounter"), None)
        if duc is not None:
            card["defaultUseCounter"] = duc
        dmc = as_int(row.get("defaultFinalMutationCountdown"), None)
        if dmc is not None:
            card["defaultFinalMutationCountdown"] = dmc
        if mutationOdds is not None:
            card["mutationOdds"] = mutationOdds

        card["finalMutation"] = {
            "outcomeWeights": {"brick": brick_w, "rewrite": rewrite_w},
            "rewritePoolDefIds": rewrite_pool,
            "brickBehavior": brick_behavior,
        }

        cards[cid] = card
    return cards


def build_mutations(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    muts: Dict[str, Any] = {}
    for row in rows:
        require_fields(row, ["id", "name", "tier"], "Mutations")
        mid = norm(row["id"])
        if mid in muts:
            die(f"Mutations: duplicate id {mid}")

        tier = norm(row["tier"])
        if tier not in list("ABCDEFGHIJ"):
            die(f"Mutations {mid}: invalid tier {tier}")

        m = {"id": mid, "name": norm(row["name"]), "tier": tier}

        add = parse_json_cell(row.get("addEffects_json"), default=None)
        if add is not None:
            if not isinstance(add, list):
                die(f"Mutations {mid}: addEffects_json must be JSON array")
            m["addEffects"] = add

        for k in ("ramCostDelta", "useCounterDelta", "finalCountdownDelta"):
            v = as_int(row.get(k), None)
            if v is not None:
                m[k] = v

        if norm(row.get("stackable")) != "":
            m["stackable"] = as_bool(row.get("stackable"), False)

        muts[mid] = m
    return muts


def build_enemies(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    enemies: Dict[str, Any] = {}
    for row in rows:
        require_fields(row, ["id", "name", "maxHP", "rotation"], "Enemies")
        eid = norm(row["id"])
        if eid in enemies:
            die(f"Enemies: duplicate id {eid}")
        enemies[eid] = {
            "id": eid,
            "name": norm(row["name"]),
            "maxHP": as_int(row["maxHP"], 1),
            "rotation": split_csv(row["rotation"]),
        }
    return enemies


def build_encounters(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    enc: Dict[str, Any] = {}
    for row in rows:
        require_fields(row, ["id", "name", "enemyIds", "weight"], "Encounters")
        eid = norm(row["id"])
        if eid in enc:
            die(f"Encounters: duplicate id {eid}")
        enc[eid] = {
            "id": eid,
            "name": norm(row["name"]),
            "enemyIds": split_csv(row["enemyIds"]),
            "weight": as_float(row["weight"], 1.0),
        }
        tags = split_csv(row.get("tags"))
        if tags:
            enc[eid]["tags"] = tags
    return enc


def build_encounter_tables(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        require_fields(row, ["act", "normal", "elite", "boss"], "EncounterTables")
        out.append({
            "act": as_int(row["act"], 1),
            "normal": split_csv(row["normal"]),
            "elite": split_csv(row["elite"]),
            "boss": split_csv(row["boss"]),
        })
    return out


def build_act_balance(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in rows:
        require_fields(row, ["act", "enemyHpMult", "enemyDmgMult", "goldNormal", "goldElite", "goldBoss"], "ActBalance")
        out.append({
            "act": as_int(row["act"], 1),
            "enemyHpMult": as_float(row["enemyHpMult"], 1.0),
            "enemyDmgMult": as_float(row["enemyDmgMult"], 1.0),
            "goldNormal": as_int(row["goldNormal"], 25),
            "goldElite": as_int(row["goldElite"], 50),
            "goldBoss": as_int(row["goldBoss"], 99),
        })
    return out


def build_relics(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for row in rows:
        require_fields(row, ["id", "name", "rarity", "description"], "Relics")
        rid = norm(row["id"])
        if rid in out:
            die(f"Relics: duplicate id {rid}")
        r = {
            "id": rid,
            "name": norm(row["name"]),
            "rarity": norm(row["rarity"]),
            "description": norm(row["description"]),
        }
        mods = parse_json_cell(row.get("mods_json"), default=None)
        if mods is not None:
            if not isinstance(mods, dict):
                die(f"Relics {rid}: mods_json must be a JSON object")
            r["mods"] = mods
        hooks = parse_json_cell(row.get("hooks_json"), default=None)
        if hooks is not None:
            if not isinstance(hooks, dict):
                die(f"Relics {rid}: hooks_json must be a JSON object")
            r["hooks"] = hooks
        out[rid] = r
    return out


def build_relic_pools(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    if not rows:
        return {"common": [], "uncommon": [], "rare": [], "boss": []}
    row = rows[0]
    return {
        "common": split_csv(row.get("common")),
        "uncommon": split_csv(row.get("uncommon")),
        "rare": split_csv(row.get("rare")),
        "boss": split_csv(row.get("boss")),
    }


def main() -> None:
    if len(sys.argv) < 3:
        die("Usage: build_gamedata.py <input.xlsx> <output.json>")

    xlsx = sys.argv[1]
    out_path = sys.argv[2]

    wb = load_workbook(xlsx, data_only=True)

    def get_sheet(name: str):
        if name not in wb.sheetnames:
            die(f"Missing sheet: {name}")
        return wb[name]

    cards = build_cards(sheet_rows_by_header(get_sheet("Cards"))[1])
    mutations = build_mutations(sheet_rows_by_header(get_sheet("Mutations"))[1])
    enemies = build_enemies(sheet_rows_by_header(get_sheet("Enemies"))[1])
    encounters = build_encounters(sheet_rows_by_header(get_sheet("Encounters"))[1])
    encounterTables = build_encounter_tables(sheet_rows_by_header(get_sheet("EncounterTables"))[1])
    actBalance = build_act_balance(sheet_rows_by_header(get_sheet("ActBalance"))[1])
    relics = build_relics(sheet_rows_by_header(get_sheet("Relics"))[1])
    relicPools = build_relic_pools(sheet_rows_by_header(get_sheet("RelicPools"))[1])

    # Cross-validation
    for eid, e in enemies.items():
        for c in e["rotation"]:
            if c not in cards:
                die(f"Enemy {eid} rotation references missing card id: {c}")

    for enc_id, enc in encounters.items():
        for enemy_id in enc["enemyIds"]:
            if enemy_id not in enemies:
                die(f"Encounter {enc_id} references missing enemy id: {enemy_id}")

    gamedata = {
        "cards": cards,
        "mutations": mutations,
        "enemies": enemies,
        "encounters": encounters,
        "encounterTables": encounterTables,
        "actBalance": actBalance,
        "relics": relics,
        "relicRewardPools": relicPools,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(gamedata, f, indent=2, ensure_ascii=False)

    print(f"[build_gamedata] Wrote {out_path} (cards={len(cards)}, muts={len(mutations)}, enemies={len(enemies)}, encounters={len(encounters)}, relics={len(relics)})")


if __name__ == "__main__":
    main()
